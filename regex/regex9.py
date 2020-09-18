# train.xlsx 데이터를 읽어들여 4개의 시트를 만들어
# 각 행의 정보를 복사하기

import re
import openpyxl

# Mr (시트명 : 남성)
# Miss (시트명 : 미혼여성)
# Mrs (시트명 : 기혼여성)
# Others (시트명 : 기타)

# 엑셀파일 로드
excel_file = openpyxl.load_workbook('./resources/train.xlsx')
sheet1 = excel_file.active

# 분류하여 저장할 엑셀파일 생성
excel_sort_file = openpyxl.Workbook()
sheet2 = excel_sort_file.active
sheet2.column_dimensions['d'].width = 70
sheet2.title = '남성'

# 두번째 시트부터는 생성
sheet3 = excel_sort_file.create_sheet(title="미혼여성")
# sheet3 = excel_sort_file.active
sheet3.column_dimensions['D'].width = 70
sheet4 = excel_sort_file.create_sheet(title="기혼여성")
# sheet4 = excel_sort_file.active
sheet4.column_dimensions['d'].width = 70
sheet5 = excel_sort_file.create_sheet(title="기타")
# sheet5 = excel_sort_file.active
sheet5.column_dimensions['d'].width = 70

print(sheet1.rows)

# 생존자수 / 사망자수 시트 생성
sheet6 = excel_sort_file.create_sheet(title='보고서')
sheet6.append(['분류', '생존자수', '사망자수', '생존율'])

# 성별 사망자수와 생존자 수 변수 선언
man_survived, man_unsurvived = 0, 0
miss_survived, miss_unsurvived = 0, 0
mrs_survived, mrs_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0

# 정규식
pattern = re.compile(r' [A-z]+\.')

for each_row in sheet1.rows:
    # 정규식과 매칭된 부분 찾기
    data = pattern.findall(each_row[3].value)
    # train.xlsx의 첫번째 행에 있는 항목 이름 복사 후 각 시트로 붙여넣기
    if each_row[0].row == 1:
        sheet2.append([each_item.value for each_item in each_row])
        # 윗줄 코드는 아래 구문의 축약형
        # for each_item in each_row:
        #   sheet2.append([each_item.value])
        sheet3.append([each_item.value for each_item in each_row])
        sheet4.append([each_item.value for each_item in each_row])
        sheet5.append([each_item.value for each_item in each_row])
    else:   # 정규식 매치된 부분 각 시트에 옮기기
        if  len(data) > 0:
            if data[0] == ' Mr.':
                sheet2.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    man_survived += 1
                else:
                    man_unsurvived += 1
            elif data[0] == ' Miss.':
                sheet3.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    miss_survived += 1
                else:
                    miss_unsurvived += 1
            elif data[0] == ' Mrs.':
                sheet4.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    mrs_survived += 1
                else:
                    mrs_unsurvived += 1
            else:
                sheet5.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    others_survived += 1
                else:
                    others_unsurvived += 1
        else:
            print(data)

# 생존율 구하기
man_survived_rate = '%.2f%%' % (
    man_survived / (man_survived + man_unsurvived) * 100
)
sheet6.append(['남성', man_survived, man_unsurvived, man_survived_rate])

miss_survived_rate = '%.2f%%' % (
    miss_survived / (miss_survived + miss_unsurvived) * 100
)
sheet6.append(['미혼여성', miss_survived, miss_unsurvived, miss_survived_rate])

mrs_survived_rate = '%.2f%%' % (
    mrs_survived / (mrs_survived + mrs_unsurvived) * 100
)
sheet6.append(['기혼여성', mrs_survived, mrs_unsurvived, mrs_survived_rate])

others_survived_rate = '%.2f%%' % (
    others_survived / (others_survived + others_unsurvived) * 100
)
sheet6.append(['기타', others_survived, others_unsurvived, others_survived_rate])


# mr_pattern = re.compile(r'Mr\.')
# miss_pattern = re.compile(r'Miss\.')
# mrs_pattern = re.compile(r'Mrs\.')

# for each_row in sheet1.rows:
#     if mr_pattern.search(each_row[3].value):
#         sheet2.append(each_row)
#     elif miss_pattern.search(each_row[3].value):
#         sheet3.append(each_row)
#     elif  mrs_pattern.search(each_row[3].value):
#         sheet4.append(each_row)
#     else:
#         sheet5.append(each_row)

excel_sort_file.save('./resources/train_sort.xlsx')
excel_sort_file.close()    
excel_file.close()