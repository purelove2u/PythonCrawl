import openpyxl

# 엑셀 파일 객체 생성
excel_file = openpyxl.Workbook()
print(excel_file.sheetnames)

# 기본으로 만들어진 시트 삭제
excel_file.remove(excel_file["Sheet"])

# 새로운 첫번째 시트 생성하기
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 두번째 시트 생성하기
sheet2 = excel_file.create_sheet(index=1, title="매출표")
# 생성된 시트 확인하기
print(excel_file.sheetnames)


excel_file.save("./resources/test1.xlsx")
