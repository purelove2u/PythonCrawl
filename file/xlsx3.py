import openpyxl
import openpyxl.utils.cell as utils
from openpyxl.styles import Alignment, Font

# 엑셀 파일 객체 생성
excel_file = openpyxl.Workbook()
print(excel_file.sheetnames)  # ['Sheet']

# 기본으로 만들어진 시트 삭제
excel_file.remove(excel_file["Sheet"])

# 새로운 첫번째 시트 생성하기
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 두번째 시트 생성하기
sheet2 = excel_file.create_sheet(index=1, title="매출표")
# 생성된 시트 확인하기
print(excel_file.sheetnames)

# Column - 셀서식 지정하기
for col in sheet1.iter_cols(min_col=1, max_col=6, min_row=1, max_row=3):
    for each_cell in col:
        # 셀에 들어갈 값 설정
        each_cell.value = utils.get_column_letter(each_cell.column)
        # 정렬 기준 설정
        each_cell.alignment = Alignment(horizontal='right', vertical='center')
        # 폰트 설정
        each_cell.font = Font(bold=True, name='Arial', size=12, underline='single', color='1bb638')

# 매출표 - 셀서식 지정하기
for row in sheet2.iter_cols(min_col=1, max_col=3, min_row=1, max_row=6):
    for each_cell in row:
        # 셀에 들어갈 값 설정
        each_cell.value = each_cell.row
        # 정렬 기준 설정
        each_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 폰트 설정
        each_cell.font = Font(italic=True, name='Consolas', size=10, color='ff0000')



excel_file.save("./resources/test2.xlsx")
